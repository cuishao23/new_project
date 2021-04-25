import logging
import xlwt
import openpyxl
import os
import multiprocessing
import time
from starter.config.common import LOGGING
from starter.dao.mysql import user_session_scope, new_session_scope
from starter.dao.user_mysql import *
from starter.dao.new_mysql import add_g100_user_infos, get_new_user_infos, get_dec_user_infos, update_user_infos, \
    get_user_info, get_idcard_info, update_move_member
from starter.utils.des_ecb import des_encrypt
from starter.utils.thread import mythread
from starter.utils.idcard import get_card_info

logger = logging.getLogger(LOGGING['loggername'])


def init_importer():
    '''
    :user原始数据导入
    '''
    logger.info('user importer ...')
    with user_session_scope() as session:
        g100_user_info_list = get_g100_user_list(session)
        with new_session_scope() as session:
            for info in g100_user_info_list:
                add_g100_user_infos(session, info)

    logger.info('user importer ok!')
    print('user importer ok!')


def encrypt_des():
    import threading
    '''
    :手机号des-ecb加解密
    '''
    import time
    t1 = time.time()

    with user_session_scope() as session:
        user_list = get_dec_user_infos(session)

        pool = multiprocessing.Pool(4)  # 创建一个线程池
        user_data = pool.map(des_encrypt, user_list)  # 往线程池中填线程
        pool.close()  # 关闭线程池，不再接受线程
        pool.join()  # 等待线程池中线程全部执行完

        for data in user_data:
            if data:
                update_user_infos(session, data)

    t2 = time.time()
    print('耗时:%4f' % (t2 - t1))
    print('encrypt_des ok!')


# def encrypt_des():
#     '''
#     :手机号des-ecb加解密
#     '''
#     import time
#     t1 = time.time()
#
#     with user_session_scope() as session:
#         user_list = get_dec_user_infos(session)
#         print(len(user_list))
#
#         pool = multiprocessing.Pool(4)    # 创建一个线程池
#         user_data = pool.map(des_encrypt, user_list)  # 往线程池中填线程
#         pool.close()  # 关闭线程池，不再接受线程
#         pool.join()  # 等待线程池中线程全部执行完
#
#         for data in user_data:
#             if data:
#                 update_user_infos(session, data)
#
#     t2 = time.time()
#     print('耗时:%4f' % (t2 - t1))
#
#     print('encrypt_des ok!')

# def encrypt_des():
#     '''
#     :手机号des-ecb加解密
#     '''
#     t1 = time.time()
#
#     with user_session_scope() as session:
#         user_list = get_dec_t_user(session)
#
#         pool = multiprocessing.Pool(4)  # 创建一个线程池
#         user_data = pool.map(des_encrypt, user_list)  # 往线程池中填线程
#         pool.close()  # 关闭线程池，不再接受线程
#         pool.join()  # 等待线程池中线程全部执行完
#
#         for data in user_data:
#             if data:
#                 update_t_user(session, data)
#
#     t2 = time.time()
#     print('耗时:%4f' % (t2 - t1))
#
#     print('encrypt_des ok!')


# def encrypt_des():
#     '''
#     :手机号des-ecb加解密
#     '''
#     import time
#     t1 = time.time()
#
#     with user_session_scope() as session:
#         user_list = get_dec_tygy_user(session)
#
#         pool = multiprocessing.Pool(4)    # 创建一个线程池
#         user_data = pool.map(des_encrypt, user_list)  # 往线程池中填线程
#         pool.close()  # 关闭线程池，不再接受线程
#         pool.join()  # 等待线程池中线程全部执行完
#
#         for data in user_data:
#             if data:
#                 update_tygy_user(session, data)
#
#     t2 = time.time()
#     print('耗时:%4f' % (t2 - t1))
#
#     print('encrypt_des ok!')

# def encrypt_des():
#     '''
#     :手机号des-ecb加解密
#     '''
#     import time
#     t1 = time.time()
#
#     with user_session_scope() as session:
#         user_list = get_spark_user(session)
#
#         pool = multiprocessing.Pool(4)    # 创建一个线程池
#         user_data = pool.map(des_encrypt, user_list)  # 往线程池中填线程
#         pool.close()  # 关闭线程池，不再接受线程
#         pool.join()  # 等待线程池中线程全部执行完
#
#         for data in user_data:
#             if data:
#                 update_spark_user(session, data)
#
#     t2 = time.time()
#     print('耗时:%4f' % (t2 - t1))
#
#     print('encrypt_des ok!')

def export_data():
    '''
    :数据导出
    :return: excel
    '''
    try:
        with user_session_scope() as session:
            user_list = get_user_info(session)
            print(user_list)

        # pandas to_excel 100万（方法1）
        import pandas as pd
        df = pd.DataFrame(user_list)
        df = df.set_index('id')
        writer = pd.ExcelWriter(r'/Users/cuixin/Desktop/data.xlsx', engine='xlsxwriter', options={'strings_to_urls': False})  # 不将字符串转换为URL的选项创建ExcelWriter对象
        df.to_excel(writer)
        writer.close()
        print('ok！')

        template = {
            'userInfo': {
                'title': (
                'id', '姓名', '手机', '创建时间', '微信登陆id', '活动代号', '登陆session', '企业id', '第三方推广应用appid', '渠道号', '当前剩余积分',
                '用户积分', '用户uid', '角色', '状态', '昵称', '照片', '性别', '国家', '省份', '城市', '证件类型', '证件号码', '详细地址', '关联用户id',
                'jd号', 'verifytime', 'real_name', 'real_idcard', 'conflict_idcard', 'phone_token', '在线时间', 'validity',
                'logindays', '邮箱', '生日', 'app对应id', '登陆方式', '登陆id', '区/县', '街道'),
                'field': (
                'id', 'name', 'mobile', 'create_time', 'openid', 'app', 'session_key', 'gid', 'extappid', 'dist',
                'point', 'total_point', 'uid', 'role', 'status', 'nickname', 'avatarurl', 'gender', 'country',
                'province', 'city', 'id_type', 'id_number', 'address', 'reg_user_id', 'jdid', 'verifytime', 'real_name',
                'real_idcard', 'conflict_idcard', 'phone_token', 'token_time', 'validity', 'logindays', 'email',
                'birth', 'appid', 'login_type', 'login_id', 'district', 'street')
            }
        }

        # xlwt_to_excel 65536 （方法2）
        # inspect_result = xlwt.Workbook()
        # sheet = inspect_result.add_sheet('userInfo')
        #
        # for c, name in enumerate(template['userInfo']['title']):
        #     sheet.write(0, c, name)
        # for r, info in enumerate(user_list):
        #     for c, name in enumerate(template['userInfo']['field']):
        #         value = info.get(name, '')
        #         sheet.write(r+1, c, str(value))
        # os.makedirs('/Users/cuixin/Desktop/data', exist_ok=True)
        # inspect_result.save('/Users/cuixin/Desktop/data/today.xls')

        # openpyxl_to_excel 100万 （方法3）
        # inspect_result = openpyxl.Workbook()
        # sheet = inspect_result.create_sheet('userInfo')
        # # sheet = inspect_result.active
        # for c, name in enumerate(template['userInfo']['title']):
        #     print(c)
        #     print(name)
        #     sheet.cell(1, c + 1, name)
        # for r, info in enumerate(user_list):
        #     for c, name in enumerate(template['userInfo']['field']):
        #         value = info.get(name, '')
        #         sheet.cell(r + 1, c + 1, str(value))
        # os.makedirs('/Users/cuixin/Desktop/data', exist_ok=True)
        # inspect_result.save('/Users/cuixin/Desktop/data/样例数据.xls')
        #
        # logger.info('export_data ok!')
        # print('export_data ok!')
    except Exception as e:
        logger.error(e)
        print('export_data error=%s' % e)


# def export_data():
#     '''
#     :数据导出
#     :return: excel
#     '''
#     try:
#         template = {
#             'userInfo': {
#                 'title': ('app', '月份', '数量', '省份'),
#                 'field': ('user_app', 'user_month', 'user_count', 'user_prince')
#             }
#         }
#         with user_session_scope() as session:
#             user_list = get_new_user_infos(session)
#
#         # 65536
#         inspect_result = xlwt.Workbook()
#         sheet = inspect_result.add_sheet('userInfo')
#
#         for c, name in enumerate(template['userInfo']['title']):
#             sheet.write(0, c, name)
#         for r, info in enumerate(user_list):
#             for c, name in enumerate(template['userInfo']['field']):
#                 value = info.get(name, '')
#                 sheet.write(r+1, c, str(value))
#         os.makedirs('/Users/cuixin/Desktop/data', exist_ok=True)
#         inspect_result.save('/Users/cuixin/Desktop/data/today.xls')
#
#         # 100万
#         # inspect_result = openpyxl.Workbook()
#         # sheet = inspect_result.create_sheet('userInfo')
#         # # sheet = inspect_result.active
#         # for c, name in enumerate(template['userInfo']['title']):
#         #     print(c)
#         #     print(name)
#         #     sheet.cell(1, c+1, name)
#         # for r, info in enumerate(user_list):
#         #     for c, name in enumerate(template['userInfo']['field']):
#         #         value = info.get(name, '')
#         #         sheet.cell(r + 1, c+1, str(value))
#         # os.makedirs('/Users/cuixin/Desktop/data', exist_ok=True)
#         # inspect_result.save('/Users/cuixin/Desktop/data/today.xls')
#
#         logger.info('export_data ok!')
#         print('export_data ok!')
#     except Exception as e:
#         logger.error(e)
#         print('export_data error=%s'%e)

def analyse_idcard():
    t1 = time.time()

    with user_session_scope() as session:
        user_list = get_idcard_info(session)

        pool = multiprocessing.Pool(multiprocessing.cpu_count())
        data = pool.map(get_card_info, user_list)
        pool.close()
        pool.join()

        print("data==%s" % data)
        for info in data:
            if info:
                update_move_member(session, info)

    t2 = time.time()
    print('耗时:%4f' % (t2 - t1))

    print('encrypt_des ok!')


def text():
    import openpyxl
    # 加载文件
    wb = openpyxl.load_workbook('/Users/cuixin/Desktop/HM零售商品牌商数据说明20210406.xlsx')
    # 获得sheet名称
    sheetNames = wb.sheetnames
    sheetName1 = sheetNames[0]
    sheet1 = wb[sheetName1]
    print(sheet1)
    for row in sheet1.rows:
        for cell in row:
            print(cell.value, end=',')
        print()

    # L = ['张三', '李四', '王五']
    # k = [11,22,18]
    # # excel中单元格为B2开始，即第2列，第2行
    # for i in range(len(L)):
    #     sheet1.cell(i + 2, 2).value = L[i]
    #     sheet1.cell(i + 2, 3).value = k[i]
    # for j in range(len(k)):
    #     sheet1.cell(j + 2, 3).value = k[j]
    # 保存数据，如果提示权限错误，需要关闭打开的excel
    wb.save('/Users/cuixin/Desktop/HM零售商品牌商数据说明20210406.xlsx')
